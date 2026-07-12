# -*- coding: utf-8 -*-
import csv
import io
import json
import math
import os
import sqlite3
import zlib
from collections import defaultdict
from datetime import datetime, timedelta
from html import escape as html_escape
from statistics import mean
from zoneinfo import ZoneInfo

from flask import Flask, Response, flash, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

APP_TITLE = "Dashboard Encuesta de Satisfacción - Campamento 5400"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
LOGO_ARAMARK_PATH = os.path.join(STATIC_DIR, "img", "logo_aramark.png")
LOGO_BHP_PATH = os.path.join(STATIC_DIR, "img", "logo_campamento_5400.png")
DB_PATH = os.environ.get("DATABASE_PATH", "/var/data/encuesta_5400.db")
if not os.path.exists(os.path.dirname(DB_PATH)):
    DB_PATH = os.path.join(BASE_DIR, "encuesta_5400.db")

EXPECTED_COLUMNS = [
    "FECHA",
    "Q1_RESPUESTA", "Q1_PUNTAJE",
    "Q2_RESPUESTA", "Q2_PUNTAJE",
    "Q3_RESPUESTA", "Q3_PUNTAJE",
    "Q4_RESPUESTA", "Q4_PUNTAJE",
    "Q5_RESPUESTA", "Q5_PUNTAJE",
    "TOTAL", "PROMEDIO", "COMENTARIOS",
]

DB_COLUMNS = [
    "fecha", "q1_respuesta", "q1_puntaje", "q2_respuesta", "q2_puntaje",
    "q3_respuesta", "q3_puntaje", "q4_respuesta", "q4_puntaje",
    "q5_respuesta", "q5_puntaje", "total", "promedio", "comentarios",
]

DIMENSION_LABELS = {
    "q1_puntaje": "Q1. Primera Impresión / Recepción",
    "q2_puntaje": "Q2. Calidad del Serv. Principal",
    "q3_puntaje": "Q3. Tiempos de Respuesta",
    "q4_puntaje": "Q4. Higiene y Presentación",
    "q5_puntaje": "Q5. Trato y Factor Humano",
}

RADAR_LABELS = ["Q1 Recepción", "Q2 Calidad", "Q3 Tiempos", "Q4 Higiene", "Q5 Trato"]
RED = "#ed1b2e"
RED_DARK = "#b60f1f"
WEEK_RULE_NOTE = (
    "Estructura semanal: lunes a domingo. "
    "La semana se publica al cerrarse el domingo y luego se presenta hacia atrás."
)
REPORT_TIMEZONE = ZoneInfo(os.environ.get("REPORT_TIMEZONE", "America/Santiago"))


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambiar-esta-clave-en-render")
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with get_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS encuestas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                q1_respuesta TEXT,
                q1_puntaje REAL,
                q2_respuesta TEXT,
                q2_puntaje REAL,
                q3_respuesta TEXT,
                q3_puntaje REAL,
                q4_respuesta TEXT,
                q4_puntaje REAL,
                q5_respuesta TEXT,
                q5_puntaje REAL,
                total REAL,
                promedio REAL,
                comentarios TEXT,
                imported_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def normalize_col(value) -> str:
    return (
        str(value or "")
        .strip()
        .upper()
        .replace("Á", "A")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ú", "U")
        .replace("Ñ", "N")
        .replace(" ", "_")
    )


def to_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        # Fecha serial de Excel.
        if 30000 <= float(value) <= 70000:
            return datetime(1899, 12, 30) + timedelta(days=float(value))
    text = str(value).strip()
    formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def text_value(value):
    if value is None:
        return ""
    return str(value).strip()


def find_header_row(ws, max_scan=30):
    expected = {normalize_col(c) for c in EXPECTED_COLUMNS}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        normalized = [normalize_col(cell) for cell in row]
        hits = sum(1 for col in normalized if col in expected)
        if hits >= 8:
            return row_idx, normalized
    return 1, [normalize_col(cell.value) for cell in ws[1]]


def read_uploaded_workbook(file_storage):
    filename = secure_filename(file_storage.filename or "archivo.xlsx")
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("El archivo debe estar en formato Excel .xlsx.")

    file_storage.stream.seek(0)
    wb = load_workbook(filename=file_storage.stream, data_only=True, read_only=True)

    selected = None
    for name in wb.sheetnames:
        low = name.lower()
        if "encuesta" in low and "satisf" in low:
            selected = name
            break
    if selected is None:
        selected = wb.sheetnames[0]

    ws = wb[selected]
    header_row, normalized_headers = find_header_row(ws)
    header_map = {name: idx for idx, name in enumerate(normalized_headers) if name}
    expected = [normalize_col(c) for c in EXPECTED_COLUMNS]
    missing = [col for col in expected if col not in header_map]
    if missing:
        raise ValueError("Faltan columnas requeridas: " + ", ".join(missing))

    records = []
    q_score_cols = ["q1_puntaje", "q2_puntaje", "q3_puntaje", "q4_puntaje", "q5_puntaje"]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not any(value is not None and str(value).strip() != "" for value in row):
            continue

        raw = {}
        for expected_name, db_name in zip(expected, DB_COLUMNS):
            idx = header_map[expected_name]
            raw[db_name] = row[idx] if idx < len(row) else None

        fecha = parse_date(raw["fecha"])
        if fecha is None:
            continue

        record = {
            "fecha": fecha.strftime("%Y-%m-%d %H:%M:%S"),
            "q1_respuesta": text_value(raw["q1_respuesta"]),
            "q2_respuesta": text_value(raw["q2_respuesta"]),
            "q3_respuesta": text_value(raw["q3_respuesta"]),
            "q4_respuesta": text_value(raw["q4_respuesta"]),
            "q5_respuesta": text_value(raw["q5_respuesta"]),
            "comentarios": text_value(raw["comentarios"]),
        }
        for col in q_score_cols:
            record[col] = to_float(raw[col])

        valid_scores = [record[col] for col in q_score_cols if record[col] is not None]
        total = to_float(raw["total"])
        promedio = to_float(raw["promedio"])
        if total is None and valid_scores:
            total = sum(valid_scores)
        if promedio is None and valid_scores:
            promedio = sum(valid_scores) / len(valid_scores)
        if promedio is None or promedio < 1 or promedio > 5:
            continue
        record["total"] = total
        record["promedio"] = promedio
        records.append(record)

    if not records:
        raise ValueError("No se encontraron registros válidos con fecha y promedio entre 1 y 5.")
    return records


def save_records(records, mode="replace"):
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        if mode == "replace":
            conn.execute("DELETE FROM encuestas")
        conn.executemany(
            """
            INSERT INTO encuestas (
                fecha, q1_respuesta, q1_puntaje, q2_respuesta, q2_puntaje,
                q3_respuesta, q3_puntaje, q4_respuesta, q4_puntaje,
                q5_respuesta, q5_puntaje, total, promedio, comentarios, imported_at
            ) VALUES (
                :fecha, :q1_respuesta, :q1_puntaje, :q2_respuesta, :q2_puntaje,
                :q3_respuesta, :q3_puntaje, :q4_respuesta, :q4_puntaje,
                :q5_respuesta, :q5_puntaje, :total, :promedio, :comentarios, :imported_at
            )
            """,
            [{**record, "imported_at": imported_at} for record in records],
        )
        conn.commit()
    return len(records)


def load_data():
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM encuestas ORDER BY fecha ASC, id ASC").fetchall()
    data = []
    for row in rows:
        item = dict(row)
        item["fecha_dt"] = parse_date(item.get("fecha"))
        if item["fecha_dt"] is not None and item.get("promedio") is not None:
            data.append(item)
    return data


def status_from_score(score):
    if score is None:
        return "Sin encuestas"
    if score >= 4.80:
        return "Mantener estándar"
    if score >= 4.70:
        return "Seguimiento preventivo"
    return "Generar plan de acción"


def status_class(text):
    text = (text or "").lower()
    if "mantener" in text:
        return "ok"
    if "preventivo" in text or "seguimiento" in text:
        return "warn"
    if "acción" in text or "accion" in text or "riesgo" in text:
        return "danger"
    return "neutral"


def pct_number(part, total):
    return 0.0 if total == 0 else round((part / total) * 100, 1)


def pct(part, total):
    return f"{pct_number(part, total):.1f}%"


CUMPLIMIENTO_TARGET = 4.5


def cumplimiento_from_score(score):
    """Cumplimiento: promedio >= 4.5 equivale a 100%. Bajo 4.5, se calcula proporcionalmente contra el umbral 4.5."""
    if score is None:
        return 0.0
    try:
        value = float(score)
    except (TypeError, ValueError):
        return 0.0
    if value >= CUMPLIMIENTO_TARGET:
        return 100.0
    return round(max(0.0, min(100.0, (value / CUMPLIMIENTO_TARGET) * 100.0)), 1)


def cumple_umbral(score):
    """Verdadero cuando la evaluación cumple el umbral operativo de nota 4.5."""
    try:
        return float(score) >= CUMPLIMIENTO_TARGET
    except (TypeError, ValueError):
        return False


def parse_filter_date(value):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def filter_data_by_dates(data, fecha_inicio=None, fecha_fin=None):
    start = parse_filter_date(fecha_inicio)
    end = parse_filter_date(fecha_fin)
    if start is None and end is None:
        return []
    out = []
    for item in data:
        dt = item.get("fecha_dt")
        if dt is None:
            continue
        d = dt.date()
        if start is not None and d < start:
            continue
        if end is not None and d > end:
            continue
        out.append(item)
    return out


def build_filter_label(fecha_inicio=None, fecha_fin=None):
    start = parse_filter_date(fecha_inicio)
    end = parse_filter_date(fecha_fin)
    if start and end:
        return f"{start:%d-%m-%Y} al {end:%d-%m-%Y}"
    if start:
        return f"Desde {start:%d-%m-%Y}"
    if end:
        return f"Hasta {end:%d-%m-%Y}"
    return "Sin rango seleccionado"


def start_of_week(dt):
    """Devuelve el lunes correspondiente a la fecha recibida."""
    return (
        dt - timedelta(days=dt.weekday())
    ).replace(hour=0, minute=0, second=0, microsecond=0)


def week_label(start_date):
    """
    Genera el nombre oficial de la semana.

    Ejemplo:
        Sem. 19 (04-05 al 10-05)
    """
    end_date = start_date + timedelta(days=6)
    iso_week = int(start_date.isocalendar().week)
    return f"Sem. {iso_week} ({start_date:%d-%m} al {end_date:%d-%m})"


def average(values):
    clean = [v for v in values if v is not None]
    return round(mean(clean), 2) if clean else None


def latest_report_week_start(reference_date=None):
    """
    Determina la semana que debe publicarse.

    - Domingo: publica la semana que termina ese mismo día.
    - Lunes a sábado: conserva la semana terminada el domingo anterior.

    La fecha se evalúa con la zona horaria de Chile.
    """
    if reference_date is None:
        current_date = datetime.now(REPORT_TIMEZONE).date()
    elif isinstance(reference_date, datetime):
        current_date = reference_date.date()
    else:
        current_date = reference_date

    current_monday = current_date - timedelta(days=current_date.weekday())

    if current_date.weekday() == 6:
        report_monday = current_monday
    else:
        report_monday = current_monday - timedelta(days=7)

    return datetime.combine(report_monday, datetime.min.time())


def latest_report_week_end(reference_date=None):
    """Devuelve el domingo de la última semana publicable."""
    return latest_report_week_start(reference_date) + timedelta(days=6)


def filter_data_through_closed_week(data, reference_date=None):
    """
    Excluye registros pertenecientes a una semana todavía abierta.

    Los registros de una nueva semana se incorporan automáticamente
    cuando llega el domingo correspondiente.
    """
    cutoff_date = latest_report_week_end(reference_date).date()

    return [
        item
        for item in data
        if item.get("fecha_dt") is not None
        and item["fecha_dt"].date() <= cutoff_date
    ]


def build_weekly(data, final_week=None):
    """
    Construye semanas correlativas hasta la última semana cerrada.

    Las semanas sin encuestas se conservan para mantener la secuencia
    Sem. 28, Sem. 27, Sem. 26, etc.
    """
    if final_week is None:
        final_week = latest_report_week_start()

    buckets = defaultdict(list)

    for item in data:
        fecha_dt = item.get("fecha_dt")
        if fecha_dt is None:
            continue

        week_start = start_of_week(fecha_dt)
        if week_start <= final_week:
            buckets[week_start].append(item)

    valid_dates = [
        item.get("fecha_dt")
        for item in data
        if item.get("fecha_dt") is not None
        and start_of_week(item["fecha_dt"]) <= final_week
    ]

    first_week = (
        start_of_week(min(valid_dates))
        if valid_dates
        else final_week
    )

    weekly = []
    cursor = first_week

    while cursor <= final_week:
        rows = buckets.get(cursor, [])

        record = {
            "week_start": cursor,
            "week_end": cursor + timedelta(days=6),
            "iso_year": int(cursor.isocalendar().year),
            "iso_week": int(cursor.isocalendar().week),
            "periodo": week_label(cursor),
            "n_encuestas": len(rows),
            "promedio_general": average(
                [row.get("promedio") for row in rows]
            ),
        }

        record["cumplimiento"] = cumplimiento_from_score(
            record["promedio_general"]
        )

        for column in DIMENSION_LABELS:
            record[column] = average(
                [row.get(column) for row in rows]
            )

        record["estado"] = (
            status_from_score(record["promedio_general"])
            if rows
            else "Sin encuestas"
        )

        weekly.append(record)
        cursor += timedelta(weeks=1)

    return weekly


def build_metrics(data, include_current_closed_week=True):
    """
    Calcula métricas usando exclusivamente semanas publicables.

    El reporte histórico termina en la última semana cerrada. En reportes
    filtrados, la secuencia termina en la última semana del rango que tenga
    datos, sin superar la última semana cerrada.
    """
    closed_data = filter_data_through_closed_week(data)

    if not closed_data:
        return {
            "has_data": False,
            "total": 0,
            "global_score": 0,
            "excelencia": 0,
            "cumplimiento": 0,
            "risk": [],
            "comments": [],
            "weekly": [],
            "recent_weeks": [],
            "dimensions": [],
            "analysis": [],
            "chart_labels": [],
            "chart_values": [],
            "radar_labels": RADAR_LABELS,
            "radar_values": [],
            "generated_at": datetime.now(REPORT_TIMEZONE).strftime("%Y-%m-%d %H:%M"),
            "global_status": "Sin datos",
            "latest_label": week_label(latest_report_week_start()),
            "week_rule_note": WEEK_RULE_NOTE,
        }

    current_closed_week = latest_report_week_start()

    if include_current_closed_week:
        final_week = current_closed_week
    else:
        final_week = min(
            current_closed_week,
            start_of_week(max(item["fecha_dt"] for item in closed_data)),
        )

    total = len(closed_data)
    global_score = average([item.get("promedio") for item in closed_data]) or 0
    cumplimiento = cumplimiento_from_score(global_score)
    promoters = sum(1 for item in closed_data if item.get("promedio", 0) >= 4.8)
    cumplen = sum(1 for item in closed_data if cumple_umbral(item.get("promedio")))
    no_cumplen = total - cumplen

    risk = [
        {
            "categoria": "Cumple estándar operativo",
            "criterio": "Promedio >= 4.5",
            "volumen": cumplen,
            "representacion": pct(cumplen, total),
            "kind": "ok",
        },
        {
            "categoria": "No cumple estándar operativo",
            "criterio": "Promedio < 4.5",
            "volumen": no_cumplen,
            "representacion": pct(no_cumplen, total),
            "kind": "danger",
        },
    ]

    comments = [
        item.get("comentarios", "")
        for item in sorted(
            closed_data,
            key=lambda row: row["fecha_dt"],
            reverse=True,
        )
        if item.get("comentarios", "").strip()
    ][:6]

    weekly = build_weekly(closed_data, final_week=final_week)

    # Gráfico: orden cronológico, desde semanas anteriores hasta la actual.
    chart_weeks = [
        row
        for row in weekly
        if row["promedio_general"] is not None
    ][-10:]

    # Tabla: comienza por la semana cerrada vigente y retrocede.
    recent = list(reversed(weekly[-10:]))

    latest_rows = [
        item
        for item in closed_data
        if start_of_week(item["fecha_dt"]) == final_week
    ]
    latest_label = week_label(final_week)

    dimensions = []
    for key, label in DIMENSION_LABELS.items():
        score = average([row.get(key) for row in latest_rows])
        dimensions.append(
            {
                "codigo": key[:2].upper(),
                "label": label,
                "score": score,
                "foco": score is not None and score < 4.70,
            }
        )

    sorted_dims = sorted(
        [dimension for dimension in dimensions if dimension["score"] is not None],
        key=lambda item: item["score"],
    )

    analysis = []

    if sorted_dims:
        worst = sorted_dims[0]
        best = sorted_dims[-1]
        analysis.append(
            f"Mejor dimensión de {latest_label}: "
            f"{best['label']} con {best['score']:.2f}."
        )

        if worst["score"] < 4.70:
            analysis.append(
                f"Foco preventivo de {latest_label}: "
                f"{worst['label']} registra {worst['score']:.2f}, "
                "bajo el umbral 4.70."
            )
        else:
            analysis.append(
                f"Sin foco crítico en {latest_label}: "
                f"la dimensión más baja es {worst['label']} "
                f"con {worst['score']:.2f}."
            )
    else:
        analysis.append(
            f"{latest_label} no registra encuestas para el análisis por dimensiones."
        )

    if no_cumplen > 0:
        analysis.append(
            f"Cumplimiento operativo histórico hasta {latest_label}: "
            f"{no_cumplen} evaluaciones están bajo 4.5."
        )

    if comments:
        analysis.append(
            "La revisión de comentarios recientes permite identificar "
            "causas operativas específicas detrás de los puntajes."
        )

    analysis.append(
        f"Cumplimiento histórico hasta {latest_label}: "
        f"{cumplimiento:.1f}% según la regla promedio >= 4.5."
    )

    return {
        "has_data": True,
        "total": total,
        "global_score": global_score,
        "global_status": status_from_score(global_score),
        "excelencia": pct_number(promoters, total),
        "cumplimiento": cumplimiento,
        "risk": risk,
        "comments": comments,
        "weekly": weekly,
        "recent_weeks": [
            {
                "periodo": row["periodo"],
                "n_encuestas": row["n_encuestas"],
                "promedio": row["promedio_general"],
                "cumplimiento": row.get("cumplimiento", 0),
                "estado": row["estado"],
            }
            for row in recent
        ],
        "dimensions": dimensions,
        "latest_label": latest_label,
        "analysis": analysis,
        "chart_labels": [row["periodo"] for row in chart_weeks],
        "chart_values": [row["promedio_general"] for row in chart_weeks],
        "radar_labels": RADAR_LABELS,
        "radar_values": [dimension["score"] or 0 for dimension in dimensions],
        "generated_at": datetime.now(REPORT_TIMEZONE).strftime("%Y-%m-%d %H:%M"),
        "week_rule_note": WEEK_RULE_NOTE,
    }


def export_csv_buffer(data):
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["id"] + DB_COLUMNS + ["imported_at"])
    writer.writeheader()
    for item in data:
        row = {key: item.get(key, "") for key in ["id"] + DB_COLUMNS + ["imported_at"]}
        row["fecha"] = item.get("fecha")
        writer.writerow(row)
    return output.getvalue().encode("utf-8-sig")


def add_summary_sheets(wb, metrics, suffix=""):
    ws2 = wb.create_sheet(f"Resumen Semanal{suffix}"[:31])
    ws2.append(["Desde", "Hasta", "Año", "Semana", "Periodo", "N° Encuestas", "Prom. P1", "Prom. P2", "Prom. P3", "Prom. P4", "Prom. P5", "Promedio General", "Cumplimiento", "Estado / Acción"])
    for row in metrics["weekly"]:
        ws2.append([
            row["week_start"].strftime("%Y-%m-%d"), row["week_end"].strftime("%Y-%m-%d"), row["iso_year"], row["iso_week"], row["periodo"], row["n_encuestas"],
            row.get("q1_puntaje"), row.get("q2_puntaje"), row.get("q3_puntaje"), row.get("q4_puntaje"), row.get("q5_puntaje"), row.get("promedio_general"), row.get("cumplimiento"), row.get("estado"),
        ])

    ws3 = wb.create_sheet(f"Cumplimiento{suffix}"[:31])
    ws3.append(["Categoría", "Criterio de cumplimiento", "Volumen", "Representación"])
    for item in metrics["risk"]:
        ws3.append([item["categoria"], item["criterio"], item["volumen"], item["representacion"]])

    ws4 = wb.create_sheet(f"Dimensiones{suffix}"[:31])
    ws4.append(["Dimensión", "Promedio", "Foco"])
    for item in metrics["dimensions"]:
        ws4.append([item["label"], item["score"], "Sí" if item["foco"] else "No"])

    ws5 = wb.create_sheet(f"Macrométricas{suffix}"[:31])
    ws5.append(["Indicador", "Valor"])
    ws5.append(["Total evaluaciones", metrics.get("total", 0)])
    ws5.append(["Satisfacción global", metrics.get("global_score", 0)])
    ws5.append(["Cumplimiento", metrics.get("cumplimiento", 0)])
    ws5.append(["Índice de excelencia", metrics.get("excelencia", 0)])
    ws5.append(["Regla de cumplimiento", "Promedio >= 4.5 equivale a 100%"])
    ws5.append(["Distribución de cumplimiento", "Cumple estándar operativo: promedio >= 4.5 / No cumple: promedio < 4.5"])
    ws5.append(["Estructura semanal", WEEK_RULE_NOTE])


def build_excel_buffer(data, metrics, filtered_data=None, filtered_metrics=None, filter_label=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Histórico"
    ws.append(["ID"] + EXPECTED_COLUMNS + ["IMPORTADO_EN"])
    for item in data:
        ws.append([
            item.get("id"), item.get("fecha"),
            item.get("q1_respuesta"), item.get("q1_puntaje"),
            item.get("q2_respuesta"), item.get("q2_puntaje"),
            item.get("q3_respuesta"), item.get("q3_puntaje"),
            item.get("q4_respuesta"), item.get("q4_puntaje"),
            item.get("q5_respuesta"), item.get("q5_puntaje"),
            item.get("total"), item.get("promedio"), item.get("comentarios"), item.get("imported_at"),
        ])

    add_summary_sheets(wb, metrics, " Hist")

    if filtered_metrics is not None:
        wsf = wb.create_sheet("Datos Filtro")
        wsf.append(["Rango seleccionado", filter_label or ""])
        wsf.append([])
        wsf.append(["ID"] + EXPECTED_COLUMNS + ["IMPORTADO_EN"])
        for item in filtered_data or []:
            wsf.append([
                item.get("id"), item.get("fecha"),
                item.get("q1_respuesta"), item.get("q1_puntaje"),
                item.get("q2_respuesta"), item.get("q2_puntaje"),
                item.get("q3_respuesta"), item.get("q3_puntaje"),
                item.get("q4_respuesta"), item.get("q4_puntaje"),
                item.get("q5_respuesta"), item.get("q5_puntaje"),
                item.get("total"), item.get("promedio"), item.get("comentarios"), item.get("imported_at"),
            ])
        add_summary_sheets(wb, filtered_metrics, " Filtro")

    for wsx in wb.worksheets:
        for column_cells in wsx.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            wsx.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 48)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def build_template_buffer():
    wb = Workbook()
    ws = wb.active
    ws.title = "Encuesta de Satisfacción"
    ws.append(EXPECTED_COLUMNS)
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Muy satisfecho", 5,
        "Satisfecho", 4,
        "Muy satisfecho", 5,
        "Muy satisfecho", 5,
        "Muy satisfecho", 5,
        24, 4.8, "Comentario de ejemplo",
    ])
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 36)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# -----------------------------------------------------------------------------
# Generador PDF minimalista sin pandas, numpy, matplotlib ni reportlab.
# -----------------------------------------------------------------------------

def pdf_escape(text):
    """
    Codifica texto para strings PDF usando bytes WinAnsi/CP1252 con escapes octales.
    Esto evita que lectores PDF de navegador muestren tildes y eñes como caracteres rotos.
    Ejemplo: ó se escribe como \363 dentro del PDF, no como byte ambiguo.
    """
    replacements = {
        "•": "-",
        "–": "-",
        "—": "-",
        "“": '"',
        "”": '"',
        "‘": "'",
        "’": "'",
        "…": "...",
    }
    value = str(text or "")
    for src, dst in replacements.items():
        value = value.replace(src, dst)

    raw = value.encode("cp1252", "replace")
    escaped = []
    for byte in raw:
        # Paréntesis y backslash deben escaparse en strings literales PDF.
        # Bytes no ASCII se escriben como octal para conservar tildes/ñ.
        if byte in (0x28, 0x29, 0x5C):
            escaped.append(f"\\{byte:03o}")
        elif 32 <= byte <= 126:
            escaped.append(chr(byte))
        elif byte in (9, 10, 13):
            escaped.append(" ")
        else:
            escaped.append(f"\\{byte:03o}")
    return "".join(escaped)


def hex_to_rgb(hex_color):
    h = hex_color.strip().lstrip("#")
    return tuple(int(h[i:i+2], 16) / 255 for i in (0, 2, 4))


def wrap_text(text, max_chars):
    words = str(text or "").split()
    lines = []
    current = ""
    for word in words:
        if len(current) + len(word) + 1 <= max_chars:
            current = (current + " " + word).strip()
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines or [""]


def parse_png_rgb(path):
    with open(path, "rb") as fh:
        data = fh.read()
    if data[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError("No es PNG")
    pos = 8
    width = height = bit_depth = color_type = None
    compressed = b""
    while pos < len(data):
        length = int.from_bytes(data[pos:pos+4], "big")
        chunk_type = data[pos+4:pos+8]
        chunk_data = data[pos+8:pos+8+length]
        pos += 12 + length
        if chunk_type == b"IHDR":
            width = int.from_bytes(chunk_data[0:4], "big")
            height = int.from_bytes(chunk_data[4:8], "big")
            bit_depth = chunk_data[8]
            color_type = chunk_data[9]
        elif chunk_type == b"IDAT":
            compressed += chunk_data
        elif chunk_type == b"IEND":
            break
    if bit_depth != 8 or color_type not in (2, 6):
        raise ValueError("PNG no soportado para incrustación")
    raw = zlib.decompress(compressed)
    bpp = 4 if color_type == 6 else 3
    stride = width * bpp
    rows = []
    i = 0
    prev = bytearray(stride)

    def paeth(a, b, c):
        p = a + b - c
        pa = abs(p - a)
        pb = abs(p - b)
        pc = abs(p - c)
        if pa <= pb and pa <= pc:
            return a
        if pb <= pc:
            return b
        return c

    for _ in range(height):
        filt = raw[i]
        i += 1
        scan = bytearray(raw[i:i+stride])
        i += stride
        recon = bytearray(stride)
        for x in range(stride):
            left = recon[x-bpp] if x >= bpp else 0
            up = prev[x]
            up_left = prev[x-bpp] if x >= bpp else 0
            val = scan[x]
            if filt == 0:
                recon[x] = val
            elif filt == 1:
                recon[x] = (val + left) & 0xFF
            elif filt == 2:
                recon[x] = (val + up) & 0xFF
            elif filt == 3:
                recon[x] = (val + ((left + up) // 2)) & 0xFF
            elif filt == 4:
                recon[x] = (val + paeth(left, up, up_left)) & 0xFF
            else:
                raise ValueError("Filtro PNG no soportado")
        rows.append(recon)
        prev = recon
    rgb = bytearray()
    if color_type == 6:
        for row in rows:
            for x in range(0, len(row), 4):
                r, g, b, a = row[x], row[x+1], row[x+2], row[x+3]
                alpha = a / 255
                # Se aplana sobre fondo blanco para mantener compatibilidad PDF simple.
                rgb.extend([int(r * alpha + 255 * (1 - alpha)), int(g * alpha + 255 * (1 - alpha)), int(b * alpha + 255 * (1 - alpha))])
    else:
        for row in rows:
            rgb.extend(row)
    return width, height, bytes(rgb)


def build_winansi_tounicode_stream():
    """CMap ToUnicode para que copiar/visualizar tildes en PDF sea estable."""
    pairs = []
    for code in range(32, 256):
        try:
            char = bytes([code]).decode("cp1252")
        except UnicodeDecodeError:
            continue
        if not char:
            continue
        pairs.append((code, ord(char)))

    lines = [
        "/CIDInit /ProcSet findresource begin",
        "12 dict begin",
        "begincmap",
        "/CIDSystemInfo << /Registry (Adobe) /Ordering (UCS) /Supplement 0 >> def",
        "/CMapName /WinAnsiToUnicode def",
        "/CMapType 2 def",
        "1 begincodespacerange",
        "<00> <FF>",
        "endcodespacerange",
    ]
    for start in range(0, len(pairs), 100):
        chunk = pairs[start:start + 100]
        lines.append(f"{len(chunk)} beginbfchar")
        for code, uni in chunk:
            lines.append(f"<{code:02X}> <{uni:04X}>")
        lines.append("endbfchar")
    lines.extend([
        "endcmap",
        "CMapName currentdict /CMap defineresource pop",
        "end",
        "end",
    ])
    stream = "\n".join(lines).encode("ascii")
    return f"<< /Length {len(stream)} >>\nstream\n".encode("latin-1") + stream + b"\nendstream"


class SimplePDF:
    def __init__(self):
        self.objects = [None]
        self.catalog_id = self.reserve()
        self.pages_id = self.reserve()
        self.tounicode_id = self.add_object(build_winansi_tounicode_stream())
        self.font_id = self.add_object(
            f"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding /ToUnicode {self.tounicode_id} 0 R >>"
        )
        self.font_bold_id = self.add_object(
            f"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding /ToUnicode {self.tounicode_id} 0 R >>"
        )
        self.pages = []
        self.images = {}

    def reserve(self):
        self.objects.append(None)
        return len(self.objects) - 1

    def set_object(self, obj_id, data):
        self.objects[obj_id] = data if isinstance(data, bytes) else data.encode("latin-1")

    def add_object(self, data):
        self.objects.append(data if isinstance(data, bytes) else data.encode("latin-1"))
        return len(self.objects) - 1

    def add_image(self, name, path):
        try:
            width, height, rgb = parse_png_rgb(path)
            comp = zlib.compress(rgb)
            obj = (
                f"<< /Type /XObject /Subtype /Image /Width {width} /Height {height} "
                f"/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /FlateDecode /Length {len(comp)} >>\nstream\n"
            ).encode("latin-1") + comp + b"\nendstream"
            obj_id = self.add_object(obj)
            self.images[name] = {"id": obj_id, "width": width, "height": height}
        except Exception:
            pass

    def add_page(self, commands):
        content_bytes = "\n".join(commands).encode("latin-1")
        content_id = self.add_object(f"<< /Length {len(content_bytes)} >>\nstream\n".encode("latin-1") + content_bytes + b"\nendstream")
        page_id = self.reserve()
        xobjects = " ".join(f"/{name} {info['id']} 0 R" for name, info in self.images.items())
        xobject_part = f" /XObject << {xobjects} >>" if xobjects else ""
        page = (
            f"<< /Type /Page /Parent {self.pages_id} 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 {self.font_id} 0 R /F2 {self.font_bold_id} 0 R >>{xobject_part} >> "
            f"/Contents {content_id} 0 R >>"
        )
        self.set_object(page_id, page)
        self.pages.append(page_id)

    def finish(self):
        kids = " ".join(f"{pid} 0 R" for pid in self.pages)
        self.set_object(self.pages_id, f"<< /Type /Pages /Kids [ {kids} ] /Count {len(self.pages)} >>")
        self.set_object(self.catalog_id, f"<< /Type /Catalog /Pages {self.pages_id} 0 R >>")
        out = io.BytesIO()
        out.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
        offsets = [0]
        for idx in range(1, len(self.objects)):
            offsets.append(out.tell())
            out.write(f"{idx} 0 obj\n".encode("latin-1"))
            out.write(self.objects[idx] or b"")
            out.write(b"\nendobj\n")
        xref = out.tell()
        out.write(f"xref\n0 {len(self.objects)}\n".encode("latin-1"))
        out.write(b"0000000000 65535 f \n")
        for off in offsets[1:]:
            out.write(f"{off:010d} 00000 n \n".encode("latin-1"))
        out.write(f"trailer\n<< /Size {len(self.objects)} /Root {self.catalog_id} 0 R >>\nstartxref\n{xref}\n%%EOF".encode("latin-1"))
        out.seek(0)
        return out


def cmd_set_fill(color):
    r, g, b = hex_to_rgb(color)
    return f"{r:.3f} {g:.3f} {b:.3f} rg"


def cmd_set_stroke(color):
    r, g, b = hex_to_rgb(color)
    return f"{r:.3f} {g:.3f} {b:.3f} RG"


def rect(x, y, w, h, color):
    return [cmd_set_fill(color), f"{x:.1f} {y:.1f} {w:.1f} {h:.1f} re f"]


def line(x1, y1, x2, y2, color="#dddddd", width=1):
    return [cmd_set_stroke(color), f"{width:.1f} w", f"{x1:.1f} {y1:.1f} m {x2:.1f} {y2:.1f} l S"]


def text(x, y, value, size=10, color="#303030", bold=False):
    r, g, b = hex_to_rgb(color)
    font = "/F2" if bold else "/F1"
    return f"BT {r:.3f} {g:.3f} {b:.3f} rg {font} {size:.1f} Tf {x:.1f} {y:.1f} Td ({pdf_escape(value)}) Tj ET"


def image_cmd(name, x, y, w, h):
    return f"q {w:.1f} 0 0 {h:.1f} {x:.1f} {y:.1f} cm /{name} Do Q"


def pdf_header(cmds, pdf):
    if "LogoA" in pdf.images:
        cmds.append(image_cmd("LogoA", 42, 790, 115, 29))
    else:
        cmds.append(text(42, 804, "aramark", 16, RED, True))
    if "LogoB" in pdf.images:
        cmds.append(image_cmd("LogoB", 345, 792, 185, 16.5))
    else:
        cmds.append(text(360, 804, "ESCONDIDA | BHP", 15, "#f58220", True))
    cmds.extend(line(42, 778, 553, 778, RED, 2.5))


def draw_section(cmds, y, title):
    cmds.extend(rect(42, y - 4, 511, 24, RED))
    cmds.append(text(50, y + 3, title, 11, "#ffffff", True))
    return y - 36


def draw_kpi_cards(cmds, y, metrics):
    cards = [
        ("TOTAL DE EVALUACIONES", str(metrics["total"]), "Muestra analizada"),
        ("SATISFACCIÓN GLOBAL", f"{metrics['global_score']:.2f} / 5.0", metrics["global_status"]),
        ("CUMPLIMIENTO", f"{metrics['cumplimiento']:.1f}%", "Promedio >= 4.5 = 100%"),
        ("ÍNDICE DE EXCELENCIA", f"{metrics['excelencia']:.1f}%", "Usuarios Promotores"),
    ]
    x = 42
    w = 120
    gap = 10
    for label, value, note in cards:
        cmds.extend(rect(x, y - 72, w, 62, "#fff8f9"))
        cmds.extend(line(x, y - 10, x + w, y - 10, RED, 2))
        cmds.append(text(x + 7, y - 29, label, 6.6, "#666666", True))
        value_size = 15 if len(value) > 8 else 17
        cmds.append(text(x + 16, y - 52, value, value_size, "#222222", True))
        cmds.append(text(x + 7, y - 66, note, 5.8, "#2f7d32", True))
        x += w + gap
    return y - 100

def draw_table(cmds, x, y, headers, rows, col_widths, row_h=18, max_rows=None):
    max_rows = max_rows or len(rows)
    total_w = sum(col_widths)
    cmds.extend(rect(x, y - row_h, total_w, row_h, "#f1f1f1"))
    cx = x
    for header, cw in zip(headers, col_widths):
        cmds.append(text(cx + 4, y - 13, header, 7.2, "#444444", True))
        cx += cw
    yy = y - row_h
    for row in rows[:max_rows]:
        yy -= row_h
        cmds.extend(line(x, yy + row_h, x + total_w, yy + row_h, "#dddddd", 0.4))
        cx = x
        for val, cw in zip(row, col_widths):
            shown = str(val or "")
            if len(shown) > int(cw / 4.4):
                shown = shown[: int(cw / 4.4) - 1] + "…"
            cmds.append(text(cx + 4, yy + 5, shown, 7.4, "#303030", False))
            cx += cw
    cmds.extend(line(x, yy, x + total_w, yy, "#dddddd", 0.4))
    return yy - 28


def draw_comments_and_analysis(cmds, y, metrics):
    cmds.append(text(42, y, "Comentarios Recientes Destacados", 10.5, RED, True))
    cmds.append(text(308, y, "Análisis Cualitativo", 10.5, RED, True))
    y0 = y - 18
    left_y = y0
    for comment in metrics["comments"][:4]:
        cmds.extend(line(42, left_y + 8, 42, left_y - 34, RED, 2.5))
        lines = wrap_text(comment, 45)[:3]
        ty = left_y
        for ln in lines:
            cmds.append(text(50, ty, f"\"{ln}", 7.6, "#555555"))
            ty -= 10
        left_y -= 46
    if not metrics["comments"]:
        cmds.append(text(42, left_y, "Sin comentarios registrados.", 8, "#777777"))
    right_y = y0
    for item in metrics["analysis"][:5]:
        cmds.extend(rect(308, right_y - 2, 3, 3, RED))
        for ln in wrap_text(item, 52)[:3]:
            cmds.append(text(316, right_y, ln, 7.6, "#303030"))
            right_y -= 10
        right_y -= 8
    return min(left_y, right_y) - 4


def draw_trend_chart(cmds, x, y, w, h, labels, values):
    # Eje Y fijo de 1 a 5 con separaciones claras entre cada nivel.
    cmds.extend(line(x, y, x, y + h, "#999999", 0.8))
    cmds.extend(line(x, y, x + w, y, "#999999", 0.8))
    for level in range(1, 6):
        gy = y + h * (level - 1) / 4
        cmds.extend(line(x, gy, x + w, gy, "#e5e5e5", 0.45))
        cmds.append(text(x - 23, gy - 2, f"{level}.0", 7, "#666666"))
    if values:
        pts = []
        for idx, val in enumerate(values):
            px = x + (w * idx / max(len(values) - 1, 1))
            py = y + ((float(val) - 1) / 4) * h
            pts.append((px, py))
        cmds.append(cmd_set_stroke(RED))
        cmds.append("2 w")
        path = f"{pts[0][0]:.1f} {pts[0][1]:.1f} m " + " ".join(f"{px:.1f} {py:.1f} l" for px, py in pts[1:]) + " S"
        cmds.append(path)
        for idx, (px, py) in enumerate(pts):
            cmds.append(cmd_set_fill(RED))
            cmds.append(f"{px:.1f} {py:.1f} 3.2 0 360 arc f")
            label_y = min(py + 12, y + h + 8)
            cmds.append(text(px - 8, label_y, f"{values[idx]:.2f}", 7, "#303030"))
            # Etiqueta de semana en eje X
            raw_label = str(labels[idx]) if idx < len(labels) else ""
            week_txt = raw_label
            if "Sem." in raw_label:
                try:
                    week_txt = "S" + raw_label.split("Sem.")[1].split("(")[0].strip()
                except Exception:
                    week_txt = raw_label[:6]
            cmds.append(text(px - 8, y - 16, week_txt, 6.5, "#666666"))
    cmds.append(text(x + w/2 - 15, y - 28, "Semana", 7, "#666666"))


def draw_radar_chart(cmds, cx, cy, radius, labels, values):
    n = len(values) or 5
    angles = [-math.pi / 2 + i * 2 * math.pi / n for i in range(n)]
    for level in range(1, 6):
        r = radius * level / 5
        pts = [(cx + math.cos(a) * r, cy + math.sin(a) * r) for a in angles]
        cmds.append(cmd_set_stroke("#dddddd"))
        cmds.append("0.5 w")
        cmds.append(f"{pts[0][0]:.1f} {pts[0][1]:.1f} m " + " ".join(f"{x:.1f} {y:.1f} l" for x, y in pts[1:]) + " h S")
    pts = [(cx + math.cos(a) * radius * (float(v) / 5), cy + math.sin(a) * radius * (float(v) / 5)) for a, v in zip(angles, values)]
    cmds.append(cmd_set_stroke(RED))
    cmds.append("1.8 w")
    if pts:
        cmds.append(f"{pts[0][0]:.1f} {pts[0][1]:.1f} m " + " ".join(f"{x:.1f} {y:.1f} l" for x, y in pts[1:]) + " h S")
    for a, label in zip(angles, labels):
        lx = cx + math.cos(a) * (radius + 20)
        ly = cy + math.sin(a) * (radius + 14)
        cmds.append(text(lx - 25, ly, label[:14], 6.5, "#555555"))


def draw_scale_legend(cmds, y):
    # Bloque visible para explicar la escala en el reporte exportado PDF.
    cmds.extend(rect(42, y - 76, 511, 76, "#fff8f9"))
    cmds.extend(line(42, y, 553, y, RED, 1.4))
    cmds.append(text(50, y - 15, "Escala de Cumplimiento y Excelencia", 10, RED, True))
    cmds.append(text(50, y - 31, "1.0 a 3.9: Riesgo operativo", 8, "#303030"))
    cmds.append(text(50, y - 44, "4.0 a 4.49: Bajo estándar", 8, "#303030"))
    cmds.append(text(290, y - 31, "4.5 a 4.79: Cumple estándar operativo", 8, "#303030"))
    cmds.append(text(290, y - 44, "4.8 a 5.0: Excelencia / Promotor", 8, "#303030"))
    cmds.append(text(50, y - 61, "Cumplimiento: promedio >= 4.5 = 100%.", 7.5, "#555555"))
    cmds.append(text(290, y - 61, "Excelencia: promedio >= 4.8.", 7.5, "#555555"))
    return y - 92


def pdf_footer(cmds):
    cmds.extend(line(42, 46, 553, 46, "#dddddd", 0.5))
    cmds.append(text(126, 30, f"Generado automáticamente a partir de datos operacionales • Evaluación Interna • {datetime.now():%Y-%m-%d}", 7, "#999999"))


def add_scale_page(pdf, report_title):
    cmds = []
    pdf_header(cmds, pdf)
    cmds.append(text(42, 742, "ESCALA DE CUMPLIMIENTO Y EXCELENCIA", 18, "#202020", True))
    cmds.append(text(42, 724, report_title, 11, RED, True))
    cmds.append(text(42, 704, "Esta escala explica la diferencia entre cumplimiento operativo e índice de excelencia.", 9, "#666666"))

    y = 670
    cmds.extend(rect(42, y - 210, 511, 210, "#fff8f9"))
    cmds.extend(line(42, y, 553, y, RED, 2.0))
    cmds.append(text(58, y - 30, "RANGO DE NOTA", 9, "#444444", True))
    cmds.append(text(190, y - 30, "CATEGORÍA", 9, "#444444", True))
    cmds.append(text(350, y - 30, "INTERPRETACIÓN", 9, "#444444", True))

    rows = [
        ("1.0 a 3.9", "Riesgo operativo", "Evaluación deficiente o crítica"),
        ("4.0 a 4.49", "Bajo estándar", "Aún no cumple el estándar operativo"),
        ("4.5 a 4.79", "Cumple estándar operativo", "Cumple, pero todavía no es excelencia"),
        ("4.8 a 5.0", "Excelencia / Promotor", "Alta satisfacción / estándar superior"),
    ]
    yy = y - 60
    for rango, categoria, interpretacion in rows:
        cmds.extend(line(58, yy + 12, 535, yy + 12, "#dddddd", 0.5))
        cmds.append(text(58, yy, rango, 9, "#303030", True))
        cmds.append(text(190, yy, categoria, 9, "#303030", True))
        cmds.append(text(350, yy, interpretacion, 8.5, "#303030"))
        yy -= 36

    y2 = 390
    cmds.extend(rect(42, y2 - 105, 511, 105, "#ffffff"))
    cmds.extend(line(42, y2, 553, y2, RED, 1.5))
    cmds.append(text(58, y2 - 25, "Reglas operativas", 11, RED, True))
    cmds.append(text(58, y2 - 48, "Cumplimiento operativo: promedio >= 4.5 equivale a 100%.", 9, "#303030"))
    cmds.append(text(58, y2 - 68, "Índice de excelencia: porcentaje de evaluaciones con promedio >= 4.8.", 9, "#303030"))
    cmds.append(text(58, y2 - 88, "Una encuesta puede cumplir el estándar operativo sin ser considerada excelencia.", 8.5, "#666666"))

    pdf_footer(cmds)
    pdf.add_page(cmds)


def add_report_pages(pdf, metrics, report_title, subtitle=None):
    cmds = []
    pdf_header(cmds, pdf)
    cmds.append(text(42, 742, "REPORTE EJECUTIVO DE CALIDAD", 18, "#202020", True))
    cmds.append(text(42, 724, report_title, 11, RED, True))
    if subtitle:
        cmds.append(text(42, 708, subtitle, 8.5, "#666666"))
    y = draw_section(cmds, 680, "1. Resumen Ejecutivo (Macrométricas)")
    y = draw_kpi_cards(cmds, y, metrics)
    y = draw_scale_legend(cmds, y)
    y -= 2
    y = draw_section(cmds, y, "2. Distribución del Cumplimiento Operativo")
    risk_rows = [[r["categoria"], r["criterio"], r["volumen"], r["representacion"]] for r in metrics["risk"]]
    y = draw_table(cmds, 42, y, ["CATEGORÍA", "CRITERIO", "VOLUMEN", "REP."], risk_rows, [235, 125, 60, 90], row_h=22)
    y -= 8
    y = draw_section(cmds, y, "3. La Voz del Usuario y Observaciones")
    draw_comments_and_analysis(cmds, y, metrics)
    pdf_footer(cmds)
    pdf.add_page(cmds)

    add_scale_page(pdf, report_title)

    cmds = []
    pdf_header(cmds, pdf)
    cmds.append(text(42, 742, "ANÁLISIS DE TENDENCIAS Y DESGLOSE", 18, "#202020", True))
    cmds.append(text(42, 724, report_title, 11, RED, True))
    cmds.append(text(42, 708, WEEK_RULE_NOTE, 8, "#666666"))
    y = draw_section(cmds, 680, "4. Evolución Histórica / Rango Seleccionado (Últimas 10 Semanas con Datos)")
    draw_trend_chart(cmds, 70, y - 255, 455, 225, metrics["chart_labels"], metrics["chart_values"])
    y = y - 320
    y = draw_section(cmds, y, "5. Resumen Semanas Recientes")
    recent_rows = [
        [
            row["periodo"],
            row["n_encuestas"],
            "-" if row["promedio"] is None else f"{row['promedio']:.2f}",
            "-" if row["promedio"] is None else f"{row.get('cumplimiento', 0):.1f}%",
            row["estado"],
        ]
        for row in metrics["recent_weeks"]
    ]
    draw_table(cmds, 42, y, ["SEMANA", "N° ENC.", "PROM.", "CUMP.", "ESTADO / ACCIÓN"], recent_rows, [185, 62, 55, 58, 150], row_h=22)
    pdf_footer(cmds)
    pdf.add_page(cmds)

    cmds = []
    pdf_header(cmds, pdf)
    cmds.append(text(42, 742, "ANÁLISIS DE TENDENCIAS Y DESGLOSE", 18, "#202020", True))
    cmds.append(text(42, 724, report_title, 11, RED, True))
    y = draw_section(cmds, 690, f"6. Desempeño Específico por Dimensiones ({metrics['latest_label']})")
    cmds.append(text(42, y + 10, "Una calificación por debajo de 4.70 se considera foco de atención preventivo.", 8.5, "#555555"))
    cmds.append(text(42, y - 4, "Regla de cumplimiento: promedio >= 4.5 equivale a 100%.", 8.5, "#555555"))
    draw_radar_chart(cmds, 175, y - 155, 78, metrics["radar_labels"], metrics["radar_values"])
    dy = y - 60
    for item in metrics["dimensions"]:
        color = RED if item["foco"] else "#2f7d32"
        score = "-" if item["score"] is None else f"{item['score']:.2f}" + (" (Foco)" if item["foco"] else "")
        cmds.extend(rect(330, dy - 1, 4, 4, RED))
        cmds.append(text(340, dy, item["label"], 8.2, "#303030", True))
        cmds.append(text(505, dy, score, 8.2, color, True))
        dy -= 20
    pdf_footer(cmds)
    pdf.add_page(cmds)


def generate_pdf_report(metrics, filtered_metrics=None, filter_label=None):
    pdf = SimplePDF()
    pdf.add_image("LogoA", LOGO_ARAMARK_PATH)
    pdf.add_image("LogoB", LOGO_BHP_PATH)
    add_report_pages(pdf, metrics, "REPORTE HISTÓRICO", "Evaluación de Satisfacción e Impacto del Factor Humano en la Operación")
    if filtered_metrics is not None:
        title = f"REPORTE POR FECHA SELECCIONADA: {filter_label}" if filter_label else "REPORTE POR FECHA SELECCIONADA"
        add_report_pages(pdf, filtered_metrics, title, "Comparativo del rango seleccionado conservando estructura semanal lunes-domingo")
    return pdf.finish()


@app.after_request
def force_utf8_headers(response):
    """Evita caracteres rotos en Render/navegador para tildes, ñ y símbolos."""
    if response.mimetype in {"text/html", "text/css", "text/javascript", "application/javascript", "text/csv", "application/json"}:
        response.headers["Content-Type"] = f"{response.mimetype}; charset=utf-8"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response

@app.context_processor
def inject_helpers():
    return {"json_dumps": json.dumps, "status_class": status_class}


@app.route("/")
def dashboard():
    data = load_data()
    report_data = filter_data_through_closed_week(data)
    metrics = build_metrics(report_data)
    filter_start = (request.args.get("fecha_inicio") or "").strip()
    filter_end = (request.args.get("fecha_fin") or "").strip()
    filter_active = bool(filter_start or filter_end)
    filtered_data = filter_data_by_dates(report_data, filter_start, filter_end) if filter_active else []
    filtered_metrics = build_metrics(filtered_data, include_current_closed_week=False) if filter_active else build_metrics([])
    filter_label = build_filter_label(filter_start, filter_end)
    return render_template(
        "dashboard.html",
        title=APP_TITLE,
        metrics=metrics,
        filtered_metrics=filtered_metrics,
        filter_start=filter_start,
        filter_end=filter_end,
        filter_active=filter_active,
        filter_label=filter_label,
    )


@app.route("/importar", methods=["POST"])
def importar():
    archivo = request.files.get("archivo")
    modo = request.form.get("modo", "replace")
    if not archivo or archivo.filename == "":
        flash("Debes seleccionar un archivo Excel.", "error")
        return redirect(url_for("dashboard"))
    try:
        records = read_uploaded_workbook(archivo)
        count = save_records(records, mode=modo)
        flash(f"Importación exitosa: {count} registros procesados.", "success")
    except Exception as exc:
        flash(f"Error al importar: {exc}", "error")
    return redirect(url_for("dashboard"))


@app.route("/plantilla")
def plantilla():
    buffer = build_template_buffer()
    return send_file(buffer, as_attachment=True, download_name="plantilla_encuesta_satisfaccion_5400.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/exportar/csv")
def exportar_csv():
    data = load_data()
    content = export_csv_buffer(data)
    return Response(content, content_type="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=encuesta_satisfaccion_5400.csv"})


@app.route("/exportar/excel")
def exportar_excel():
    data = load_data()
    report_data = filter_data_through_closed_week(data)
    metrics = build_metrics(report_data)
    filter_start = (request.args.get("fecha_inicio") or "").strip()
    filter_end = (request.args.get("fecha_fin") or "").strip()
    filter_active = bool(filter_start or filter_end)
    filtered_data = filter_data_by_dates(report_data, filter_start, filter_end) if filter_active else []
    filtered_metrics = build_metrics(filtered_data, include_current_closed_week=False) if filter_active else None
    buffer = build_excel_buffer(report_data, metrics, filtered_data if filter_active else None, filtered_metrics, build_filter_label(filter_start, filter_end) if filter_active else None)
    return send_file(buffer, as_attachment=True, download_name="reporte_encuesta_satisfaccion_5400.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/exportar/pdf")
def exportar_pdf():
    data = load_data()
    report_data = filter_data_through_closed_week(data)
    metrics = build_metrics(report_data)

    if not metrics["has_data"]:
        flash("No hay datos de semanas cerradas para exportar.", "error")
        return redirect(url_for("dashboard"))

    filter_start = (request.args.get("fecha_inicio") or "").strip()
    filter_end = (request.args.get("fecha_fin") or "").strip()
    filter_active = bool(filter_start or filter_end)

    filtered_data = (
        filter_data_by_dates(report_data, filter_start, filter_end)
        if filter_active
        else []
    )
    filtered_metrics = (
        build_metrics(filtered_data, include_current_closed_week=False)
        if filter_active
        else None
    )

    buffer = generate_pdf_report(
        metrics,
        filtered_metrics,
        build_filter_label(filter_start, filter_end)
        if filter_active
        else None,
    )

    return send_file(
        buffer,
        as_attachment=True,
        download_name="reporte_encuesta_satisfaccion_5400.pdf",
        mimetype="application/pdf",
    )


@app.route("/health")
def health():
    return {"status": "ok", "app": APP_TITLE}


init_db()

if __name__ == "__main__":
    app.run(debug=True)
